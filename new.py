import openpyxl
from reportlab.pdfgen import canvas
from PIL import Image
import os
import streamlit as st
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import zipfile
import io

# Define Fonts
pdfmetrics.registerFont(TTFont('Aptos', 'fonts/Aptos.ttf'))
pdfmetrics.registerFont(TTFont('Aptos-Bold', 'fonts/Aptos-Bold.ttf'))
pdfmetrics.registerFont(TTFont('Aptos-ExtraBold', 'fonts/Aptos-ExtraBold.ttf'))

# Setup for Streamlit interface
st.title("Profile Generator")
st.markdown("Upload an Excel file to Generate Template")

# File uploader for Excel file
excel_file = st.file_uploader("Upload Excel File", type=["xlsx"])

# Check if the file is uploaded
if excel_file:
    # Load Excel data
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Template image path
    template_image = 'template_kosong.png'
    template = Image.open(template_image)

    positions = {
        'name': {'position': (410, 1420), 'max_chars': 50, 'font_size': 40, 'line_height': 40, 'font_style': 'extra_bold'},
        'jabatan': {'position': (410, 1378), 'max_chars': 80, 'font_size': 40, 'line_height': 36, 'font_style': 'normal'},
        'education': {'position': (400, 1125), 'max_chars': 100, 'font_size': 28, 'line_height': 36, 'font_style': 'normal'},
        'skills': {'position': (120, 950), 'max_chars': 45, 'font_size': 26, 'line_height': 36, 'font_style': 'normal'},
        'performance': {'position': (720, 975), 'max_chars': 14, 'font_size': 28, 'line_height': 34, 'font_style': 'normal'},
        'experience': {'position': (940, 975), 'max_chars': 70, 'font_size': 25, 'line_height': 36, 'font_style': 'normal'},
        'leadership': {'position': (1800, 940), 'max_chars': 30, 'font_size': 30, 'line_height': 38, 'font_style': 'normal'},
        'qualification': {'position': (1855, 232), 'max_chars': 30, 'font_size': 27, 'line_height': 38, 'font_style': 'bold'},
    }

    output_dir = 'results'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    def get_font_style(cell):
        if cell.font.bold:
            return 'bold'
        else:
            return 'normal'

    # Setup text drawing with customizable font styles
    def draw_multiline_text(c, text, x, y, font_size=28, line_height=38, max_chars=50, font_style='normal', color=None):
        font = 'Aptos'

        if font_style == 'bold':
            font = 'Aptos-Bold'
        elif font_style == 'extra_bold':
            font = 'Aptos-ExtraBold'

        c.setFont(font, font_size)

        if color:
            c.setFillColor(color)
        else:
            c.setFillColor('black')

        lines = text.split('\n')

        final_lines = []
        for line in lines:
            while len(line) > max_chars:
                final_lines.append(line[:max_chars])
                line = line[max_chars:]
            final_lines.append(line)

        for i, line in enumerate(final_lines):
            c.drawString(x, y - (i * line_height), line)

    # Create a button to trigger PDF generation and zipping
    if st.button('Generate and Download All PDFs'):
        # List to hold all generated PDFs for zipping
        pdf_files = []

        # Generate PDFs
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            output_pdf = os.path.join(output_dir, f"output_{row[0].value}.pdf")

            # canvas size
            c = canvas.Canvas(output_pdf, pagesize=(2667, 1500))

            c.drawImage(template_image, 0, 0, width=2667, height=1500)

            for col, col_data in positions.items():
                cell = row[list(positions).index(col) + 1]
                text = str(cell.value) if cell.value is not None else ''

                if text.strip() == '':
                    continue

                x, y = col_data['position']
                max_chars = col_data['max_chars']
                font_size = col_data['font_size']
                line_height = col_data['line_height']

                font_style = col_data.get('font_style', 'normal')

                color = 'white' if col == 'education' or col == 'qualification' else None

                if col == 'experience':
                    lines = text.split('\n')
                    for i, line in enumerate(lines):
                        if (i // 2) % 2 == 0:
                            font_style = 'bold'
                        else:
                            font_style = 'normal'
                        draw_multiline_text(c, line, x, y - (i * line_height), font_size=font_size, line_height=line_height, max_chars=max_chars, font_style=font_style, color=color)
                else:
                    draw_multiline_text(c, text, x, y, font_size=font_size, line_height=line_height, max_chars=max_chars, font_style=font_style, color=color)

            c.save()

            # Add the generated PDF to the list of files
            pdf_files.append(output_pdf)

        # Create a ZIP file containing all the PDFs
        zip_filename = "generated_pdfs.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for pdf_file in pdf_files:
                zipf.write(pdf_file, os.path.basename(pdf_file))

        # Provide the ZIP file for download
        with open(zip_filename, "rb") as zip_file:
            st.download_button(
                label="Download All PDFs",
                data=zip_file,
                file_name=zip_filename,
                mime="application/zip"
            )

        st.success("All PDFs generated and zipped successfully!")
else:
    st.warning("Please upload an Excel file to get started.")
